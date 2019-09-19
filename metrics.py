import openpyxl
import datetime

#Search files
Excelsheet1 = 'u_cat_var_log_4002.xlsx'
Excelsheet2 = 'Results.xlsx'

#Open files
df1 = openpyxl.load_workbook(Excelsheet1)
results = openpyxl.load_workbook(Excelsheet2)

#Get sheet of data frames
first_sheet = df1.get_sheet_by_name('Page 1')

#Get sheet of results
results_sheet = results.get_sheet_by_name('Sheet2')

#df1 = df1[['Developer', 'Project Urgency', 'Date Developer Assigned', 'Project Complexity']]
devs = ["lucas.javier.tuozzo", "r.uribe.palencia", "sofia.fagnilli", "matias.rodriguez", "rocio.del.c.manes", "romina.ayelen.busto", "a.molero.rincon", "andres.e.ordaz.abreu", "santiago.arturo.sanz", "natalia.dourado"]

writeRow = 1
writeColumn = 1
devNum = 1

#Recorrer la lista por cada dev
for dev in devs:
    
    total = 0
    totSimple = 0
    totComplex = 0

    for row in range(1, first_sheet.max_row + 1):

        #Si año es menor o mes es menor, actualizar mes actual y guardar resultados.

        if str(first_sheet.cell(row,1).value) == dev:
            total += 1
            if str(first_sheet.cell(row,4).value) == "Simple":
                totSimple += 1
            elif str(first_sheet.cell(row,4).value) == "Complex":
                totComplex += 1
        else:
            pass
        
    resultsList = [dev,total,totSimple,totComplex]

    for i in range(1, 5):
        results_sheet.cell(devNum,i).value = resultsList[i - 1]

    devNum += 1

results.save('Results.xlsx')